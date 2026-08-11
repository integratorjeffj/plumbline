"""Generate the synthetic Voltage Systems Inc. pricing workbook.

Vendor B of the flagship scenario (docs/demo-scenario.md): lowest submitted
price, but excludes the lighting fixture allowance and electrical permit fees,
and leaves feeder/branch-circuit testing ambiguous. Those planted gaps are what
turn Voltage from cheapest-submitted into most-expensive-adjusted.

Run: python scripts/generate_voltage_xlsx.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = REPO_ROOT / "sample-data" / "bids" / "voltage_systems_pricing.xlsx"

NAVY = "FF1C2B3A"
HEADER_FILL = PatternFill("solid", fgColor="FFEEF0F2")
THIN = Side(style="thin", color="FFC7CED5")
BORDER = Border(bottom=THIN)

LINE_ITEMS = [
    ("Mobilization, supervision, and project management", 12800.00),
    ("Branch power rough-in - all areas per drawings", 61200.00),
    ("Feeders and distribution equipment", 42000.00),
    ("Lighting branch circuitry", 34500.00),
    ("Fire alarm device connections (electrical drawings)", 9400.00),
    ("Temporary power for Voltage Systems work", 4900.00),
    ("Closeout documentation and as-built markups", 2600.00),
]

EXCLUSIONS = [
    "Lighting fixtures - NOT INCLUDED. No fixture allowance carried in this proposal.",
    "Electrical permit fees - NOT INCLUDED. Assumed carried by Crestmark.",
    "Performance and payment bond - NOT INCLUDED.",
    "Utility-company charges, fees, and service connections - NOT INCLUDED.",
    "Structured cabling / Division 27 systems - NOT INCLUDED.",
    "Security and access control / Division 28 systems - NOT INCLUDED.",
]

NOTES = [
    "Pricing based on electrical Contract Drawings Revision 3 and the Project Manual issued for bid.",
    "Testing limited to standard continuity checks. Commissioning scope to be confirmed with Crestmark.",
    "Proposal valid 30 calendar days from the date above.",
    "Anticipated duration 16 weeks from mobilization, subject to approved project schedule.",
]


def _write_header(ws, row: int, values: list[str], widths: list[int] | None = None):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = HEADER_FILL
        cell.border = BORDER
    if widths:
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width


def build():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ---------- Sheet 1: Bid Summary ----------
    ws = wb.active
    ws.title = "Bid Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 52

    ws["A1"] = "VOLTAGE SYSTEMS INC."
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = "Electrical Construction - Marietta, GA"
    ws["A2"].font = Font(italic=True, size=9)

    summary = [
        ("Proposal Date:", "August 7, 2026"),
        ("Submitted To:", "Crestmark Construction Partners"),
        ("Attention:", "Daniel Cho, Senior Estimator"),
        ("Project:", "Falcon Medical Center Expansion"),
        ("Project Number:", "26-0147"),
        ("Bid Package:", "26-0147-BP-26 - Division 26 Electrical"),
        ("Drawing Revision:", "Rev 3"),
        ("Base Bid:", 167400.00),
        ("Bid Validity:", "30 calendar days"),
    ]
    for i, (label, value) in enumerate(summary, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(bold=True, size=12)

    # ---------- Sheet 2: Pricing Detail ----------
    ws2 = wb.create_sheet("Pricing Detail")
    _write_header(ws2, 1, ["Line Item", "Amount"], widths=[58, 16])
    row = 2
    for description, amount in LINE_ITEMS:
        ws2.cell(row=row, column=1, value=description)
        amount_cell = ws2.cell(row=row, column=2, value=amount)
        amount_cell.number_format = '"$"#,##0.00'
        row += 1

    total_label = ws2.cell(row=row, column=1, value="TOTAL BASE BID")
    total_label.font = Font(bold=True, color=NAVY)
    total_cell = ws2.cell(row=row, column=2, value=sum(a for _, a in LINE_ITEMS))
    total_cell.font = Font(bold=True, color=NAVY)
    total_cell.number_format = '"$"#,##0.00'

    # ---------- Sheet 3: Exclusions ----------
    ws3 = wb.create_sheet("Exclusions")
    _write_header(ws3, 1, ["Excluded Scope"], widths=[95])
    for i, exclusion in enumerate(EXCLUSIONS, start=2):
        cell = ws3.cell(row=i, column=1, value=exclusion)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---------- Sheet 4: Notes ----------
    ws4 = wb.create_sheet("Notes")
    _write_header(ws4, 1, ["Clarifications and Assumptions"], widths=[95])
    for i, note in enumerate(NOTES, start=2):
        cell = ws4.cell(row=i, column=1, value=note)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size:,} bytes)")
    print(f"  line-item sum: ${sum(a for _, a in LINE_ITEMS):,.2f}")


if __name__ == "__main__":
    build()
