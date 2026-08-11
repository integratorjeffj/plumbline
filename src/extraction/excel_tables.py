"""Deterministic Excel extraction.

Reads a vendor pricing workbook into the same page-aware representation used
for PDFs, so downstream code (AI extraction, citations, persistence) does not
care which file format a bid arrived in.

Sheet index maps to `page_number` and sheet name to the citation section label.
"Page 2, Pricing Detail" is how an estimator would actually refer to a location
in a workbook, so the citation model needs no format-specific special case.

Numbers are rendered using the cell's own number format rather than its Python
type. openpyxl hands back a whole-dollar amount as `int`, so type-sniffing would
silently emit "167400" where the sheet plainly shows "$167,400.00" -- and the
extracted text has to match what a reviewer sees on screen for a citation to
mean anything.
"""

from pathlib import Path

from openpyxl import load_workbook

from src.extraction.pdf_text import PageText

_CURRENCY_MARKERS = ("$", "USD", "[$")


def _is_currency(number_format: str | None) -> bool:
    return bool(number_format) and any(marker in number_format for marker in _CURRENCY_MARKERS)


def _format_cell(value, number_format: str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if _is_currency(number_format):
            return f"${value:,.2f}"
        if isinstance(value, float) and value != int(value):
            return f"{value:,.2f}"
        return f"{value:,}" if abs(value) >= 1000 else str(value)
    return str(value)


def extract_sheets(xlsx_path: Path) -> list[PageText]:
    xlsx_path = Path(xlsx_path)
    workbook = load_workbook(xlsx_path, data_only=True)

    sheets: list[PageText] = []
    for index, sheet_name in enumerate(workbook.sheetnames, start=1):
        worksheet = workbook[sheet_name]
        lines = [f"[Sheet: {sheet_name}]"]
        for row in worksheet.iter_rows():
            cells = [_format_cell(cell.value, cell.number_format) for cell in row]
            if not any(cell.strip() for cell in cells):
                continue
            lines.append("  ".join(cell for cell in cells if cell.strip()))
        sheets.append(PageText(page_number=index, text="\n".join(lines)))

    workbook.close()
    return sheets


def sheet_names(xlsx_path: Path) -> list[str]:
    workbook = load_workbook(Path(xlsx_path), data_only=True)
    names = list(workbook.sheetnames)
    workbook.close()
    return names
